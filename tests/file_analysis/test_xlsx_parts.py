"""XLSX parts outside the cell grid must be pseudonymized and revealed.

Covers ``docProps/core.xml``, ``docProps/custom.xml``, cell comments, sheet
headers/footers and pivot caches — every surface that a plain openpyxl
round-trip preserves verbatim (and that a redacted workbook therefore leaked).
"""

from __future__ import annotations

import datetime
import io

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from noirdoc.file_analysis.xlsx_inference import pseudonymize_xlsx_smart
from noirdoc.file_analysis.xlsx_parts import (
    count_unsupported_part_pii,
    pseudonymize_workbook_parts,
)
from noirdoc.file_reidentification.service import reidentify_file_bytes
from noirdoc.pseudonymization.mapper import PseudonymMapper
from tests.file_analysis.xlsx_helpers import (
    SubstringDetector,
    assert_no_part_contains,
    inject_app_props,
    inject_pivot,
    inject_threaded_comment_parts,
    part_names,
    workbook_bytes,
)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Name", "Betrag"])
    ws.append(["Anna Mueller", 10])
    # openpyxl defaults creator to "openpyxl"; set it explicitly so counts are intentional.
    wb.properties.creator = "Anna Mueller"
    wb.properties.lastModifiedBy = "Dora Klein"
    return wb


# ── docProps/core.xml ────────────────────────────────────


async def test_core_author_fields_forced_to_person():
    wb = _workbook()
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    assert wb.properties.creator == "<<PERSON_1>>"
    assert wb.properties.lastModifiedBy == "<<PERSON_2>>"
    assert result.entity_count == 2
    assert result.entity_types == {"PERSON": 2}
    assert result.classifications["core.creator"] == "PERSON (forced)"
    assert result.classifications["core.lastModifiedBy"] == "PERSON (forced)"
    assert mapper.reverse_lookup("<<PERSON_1>>") == "Anna Mueller"


async def test_core_author_field_with_email_is_labelled_email():
    wb = _workbook()
    wb.properties.lastModifiedBy = "dora.klein@example.com"

    await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert wb.properties.lastModifiedBy == "<<EMAIL_1>>"


async def test_core_author_shares_placeholder_with_matching_cell_value():
    wb = _workbook()
    mapper = PseudonymMapper()
    cell_placeholder = mapper.get_or_create("Anna Mueller", "PERSON")

    await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    assert wb.properties.creator == cell_placeholder


async def test_core_free_text_fields_go_through_detector():
    wb = _workbook()
    wb.properties.title = "Kundenliste Anna Mueller"
    wb.properties.subject = "Mandant Schmidt GmbH"
    wb.properties.description = "Kontakt: anna@example.com"
    wb.properties.keywords = "Anna Mueller, Bestand"
    wb.properties.category = "Kanzlei"
    detector = SubstringDetector(
        {"Anna Mueller": "PERSON", "Schmidt GmbH": "ORGANIZATION", "anna@example.com": "EMAIL"}
    )
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    assert wb.properties.title == "Kundenliste <<PERSON_1>>"
    assert wb.properties.subject == "Mandant <<ORGANIZATION_1>>"
    assert wb.properties.description == "Kontakt: <<EMAIL_1>>"
    assert wb.properties.keywords == "<<PERSON_1>>, Bestand"
    assert wb.properties.category == "Kanzlei"
    # creator + lastModifiedBy (forced) + title + subject + description + keywords
    assert result.entity_types == {"PERSON": 4, "ORGANIZATION": 1, "EMAIL": 1}
    assert result.classifications["core.title"] == "PERSON (detected)"


async def test_core_none_and_blank_fields_are_skipped():
    wb = _workbook()
    wb.properties.creator = None
    wb.properties.lastModifiedBy = "   "
    wb.properties.title = None

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert result.entity_count == 0
    assert wb.properties.creator is None
    assert wb.properties.lastModifiedBy == "   "


async def test_generic_creator_is_still_mapped_no_allowlist():
    wb = _workbook()
    wb.properties.creator = "Microsoft Office User"
    wb.properties.lastModifiedBy = "openpyxl"

    await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert wb.properties.creator == "<<PERSON_1>>"
    assert wb.properties.lastModifiedBy == "<<PERSON_2>>"


# ── docProps/custom.xml ──────────────────────────────────


async def test_custom_string_props_pseudonymized_other_types_untouched():
    from openpyxl.packaging.custom import (
        BoolProperty,
        DateTimeProperty,
        FloatProperty,
        IntProperty,
        LinkProperty,
        StringProperty,
    )

    wb = _workbook()
    seit = datetime.datetime(2024, 1, 1)
    for prop in (
        StringProperty(name="Mandant", value="Akte Schmidt GmbH"),
        StringProperty(name="Leer", value=None),
        IntProperty(name="Akte", value=42),
        FloatProperty(name="Quote", value=0.5),
        BoolProperty(name="Aktiv", value=True),
        DateTimeProperty(name="Seit", value=seit),
        LinkProperty(name="Link", value="Daten!A1"),
    ):
        wb.custom_doc_props.append(prop)
    detector = SubstringDetector({"Schmidt GmbH": "ORGANIZATION"})

    result = await pseudonymize_workbook_parts(wb, detector, PseudonymMapper(), "de")

    props = wb.custom_doc_props
    assert props["Mandant"].value == "Akte <<ORGANIZATION_1>>"
    assert props["Leer"].value is None
    assert props["Akte"].value == 42
    assert props["Quote"].value == 0.5
    assert props["Aktiv"].value is True
    assert props["Seit"].value == seit
    assert props["Link"].value == "Daten!A1"
    assert result.classifications["custom.Mandant"] == "ORGANIZATION (detected)"
    assert result.entity_types == {"PERSON": 2, "ORGANIZATION": 1}


# ── cell comments ────────────────────────────────────────


async def test_comment_author_forced_and_text_detected():
    wb = _workbook()
    ws = wb["Daten"]
    ws["A2"].comment = Comment("Kunde anna@example.com", "Dora Klein")
    detector = SubstringDetector({"anna@example.com": "EMAIL"})
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    comment = ws["A2"].comment
    assert comment.author == "<<PERSON_2>>"  # same placeholder as lastModifiedBy "Dora Klein"
    assert comment.text == "Kunde <<EMAIL_1>>"
    assert result.classifications["comment.author@sheet1!A2"] == "PERSON (forced)"
    assert result.classifications["comment.text@sheet1!A2"] == "EMAIL (detected)"
    assert result.entity_types == {"PERSON": 3, "EMAIL": 1}


async def test_comment_author_prefix_in_text_replaced_without_detector():
    wb = _workbook()
    ws = wb["Daten"]
    # Excel's legacy comment convention: the body starts with "Author:" — NER is not
    # reliable on a bare name followed by a colon, so this must be deterministic.
    ws["A2"].comment = Comment("Dora Klein:\nbitte prüfen", "Dora Klein")

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert ws["A2"].comment.text == "<<PERSON_2>>:\nbitte prüfen"
    assert result.entity_types == {"PERSON": 4}


async def test_empty_comment_is_tolerated():
    wb = _workbook()
    wb["Daten"]["A2"].comment = Comment("", "")

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert result.entity_count == 2  # only creator + lastModifiedBy


async def test_comment_parts_hold_no_original_after_save():
    wb = _workbook()
    ws = wb["Daten"]
    ws["A2"].comment = Comment("Dora Klein:\nKontakt anna@example.com", "Dora Klein")
    ws["A2"].value = "<<PERSON_1>>"  # the cell grid is not this module's job
    detector = SubstringDetector({"anna@example.com": "EMAIL"})

    await pseudonymize_workbook_parts(wb, detector, PseudonymMapper(), "de")

    out = workbook_bytes(wb)
    assert any("comments" in n for n in part_names(out))
    assert any(n.endswith(".vml") for n in part_names(out))
    assert_no_part_contains(out, ["Anna Mueller", "Dora Klein", "anna@example.com"])


# ── sheet headers / footers ──────────────────────────────


async def test_header_footer_parts_pseudonymized():
    wb = _workbook()
    ws = wb["Daten"]
    ws.oddHeader.left.text = "Erstellt von Anna Mueller"
    ws.oddFooter.center.text = "Kontakt anna@example.com"
    ws.evenHeader.right.text = "Anna Mueller"
    ws.evenFooter.left.text = "Seite &P"
    ws.firstHeader.center.text = "Anna Mueller"
    ws.firstFooter.right.text = "Anna Mueller"
    detector = SubstringDetector({"Anna Mueller": "PERSON", "anna@example.com": "EMAIL"})

    result = await pseudonymize_workbook_parts(wb, detector, PseudonymMapper(), "de")

    assert ws.oddHeader.left.text == "Erstellt von <<PERSON_1>>"
    assert ws.oddFooter.center.text == "Kontakt <<EMAIL_1>>"
    assert ws.evenHeader.right.text == "<<PERSON_1>>"
    assert ws.evenFooter.left.text == "Seite &P"
    assert ws.firstHeader.center.text == "<<PERSON_1>>"
    assert ws.firstFooter.right.text == "<<PERSON_1>>"
    assert result.classifications["header.odd.left@sheet1"] == "PERSON (detected)"
    assert result.classifications["footer.odd.center@sheet1"] == "EMAIL (detected)"
    assert result.entity_types == {"PERSON": 6, "EMAIL": 1}
    ws["A2"].value = "<<PERSON_1>>"
    assert_no_part_contains(workbook_bytes(wb), ["Anna Mueller", "anna@example.com"])


# ── pivot caches ─────────────────────────────────────────

_PIVOT_ROWS = [["Anna Mueller", 10], ["Ben Schulz", 20]]


def _pivot_workbook(
    *,
    field_name: str = "Name",
    shared: list[str] | None = None,
    records: list[list[tuple[str, object]]] | None = None,
    tables: int = 1,
    captions: dict[str, str] | None = None,
    group_items: list[str] | None = None,
    table_sheets: list[str] | None = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append([field_name, "Betrag"])
    for row in _PIVOT_ROWS:
        ws.append(row)
    if table_sheets and "sheet2" in table_sheets:
        wb.create_sheet("Zwei")
    wb.properties.creator = "Anna Mueller"
    wb.properties.lastModifiedBy = "Dora Klein"
    xlsx = inject_pivot(
        workbook_bytes(wb),
        refreshed_by="Dora Klein",
        fields=[
            (field_name, ["Anna Mueller", "Ben Schulz"] if shared is None else shared),
            ("Betrag", None),
        ],
        records=records if records is not None else [[("x", 0), ("n", 10)], [("x", 1), ("n", 20)]],
        tables=tables,
        captions=captions,
        group_items=group_items,
        table_sheets=table_sheets,
    )
    loaded = load_workbook(io.BytesIO(xlsx))
    loaded_pivots = sum(len(ws._pivots) for ws in loaded.worksheets)
    assert loaded_pivots == tables, "fixture: pivot did not load"
    return loaded


def _shared_values(cache: object, field_index: int = 0) -> list[object]:
    return [item.v for item in cache.cacheFields[field_index].sharedItems._fields]  # type: ignore[attr-defined]


async def test_pivot_cache_scrubbed_consistently_with_sheet():
    wb = _pivot_workbook()
    mapper = PseudonymMapper()
    # Simulate the cell pass that ran before us: both names are already mapped.
    anna = mapper.get_or_create("Anna Mueller", "PERSON")
    ben = mapper.get_or_create("Ben Schulz", "PERSON")

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    cache = wb["Daten"]._pivots[0].cache
    assert _shared_values(cache) == [anna, ben]
    assert cache.refreshedBy == mapper.get_or_create("Dora Klein", "PERSON")
    records = [[item.v for item in row._fields] for row in cache.records.r]
    assert records == [[0, 10.0], [1, 20.0]]  # Index / Number items untouched
    assert result.classifications["pivot1.Name"] == "PERSON (header)"
    assert result.classifications["pivot1.refreshedBy"] == "PERSON (forced)"
    # creator + lastModifiedBy + refreshedBy + 2 shared items
    assert result.entity_types == {"PERSON": 5}

    ws = wb["Daten"]
    ws["A2"].value, ws["A3"].value = anna, ben
    out = workbook_bytes(wb)
    assert_no_part_contains(out, ["Anna Mueller", "Ben Schulz", "Dora Klein"])
    reloaded = load_workbook(io.BytesIO(out))["Daten"]
    assert len(reloaded._pivots) == 1
    assert _shared_values(reloaded._pivots[0].cache) == [anna, ben]


async def test_pivot_inline_record_text_scrubbed():
    wb = _pivot_workbook(
        shared=[], records=[[("s", "Anna Mueller"), ("n", 10)], [("s", "Ben Schulz"), ("n", 20)]]
    )
    mapper = PseudonymMapper()

    await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    cache = wb["Daten"]._pivots[0].cache
    first_column = [row._fields[0].v for row in cache.records.r]
    assert first_column == [
        mapper.get_or_create("Anna Mueller", "PERSON"),
        mapper.get_or_create("Ben Schulz", "PERSON"),
    ]
    assert [row._fields[1].v for row in cache.records.r] == [10.0, 20.0]


async def test_pivot_field_classified_by_sampling_when_header_has_no_keyword():
    wb = _pivot_workbook(field_name="Spalte1")
    detector = SubstringDetector({"Anna Mueller": "PERSON"})

    result = await pseudonymize_workbook_parts(wb, detector, PseudonymMapper(), "de")

    assert result.classifications["pivot1.Spalte1"] == "PERSON (sampled)"
    values = _shared_values(wb["Daten"]._pivots[0].cache)
    assert all(isinstance(v, str) and v.startswith("<<PERSON_") for v in values), values


async def test_pivot_unclassified_field_left_alone():
    wb = _pivot_workbook(field_name="Notiz", shared=["leave alone", "also"])

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert _shared_values(wb["Daten"]._pivots[0].cache) == ["leave alone", "also"]
    assert "pivot1.Notiz" not in result.classifications


async def test_shared_cache_not_double_mapped():
    wb = _pivot_workbook(tables=2)
    ws = wb["Daten"]
    assert ws._pivots[0].cache is ws._pivots[1].cache, "fixture: pivots must share one cache"
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    assert not any(v.startswith("<<") for v in mapper.get_mapping_summary().values())
    assert result.entity_types == {"PERSON": 5}
    assert _shared_values(ws._pivots[0].cache) == ["<<PERSON_1>>", "<<PERSON_3>>"]


# ── parts the writer drops (count-only) ──────────────────


def test_count_unsupported_part_pii_reports_app_props_and_persons():
    data = workbook_bytes(_workbook())
    data = inject_app_props(data, manager="Dora Klein", company="Schmidt GmbH")
    data = inject_threaded_comment_parts(data, display_name="Emil Roth", text="bitte prüfen")

    result = count_unsupported_part_pii(data)

    assert result.entity_types == {"PERSON": 2, "ORGANIZATION": 1}
    assert result.classifications == {
        "app.Manager": "PERSON (dropped)",
        "app.Company": "ORGANIZATION (dropped)",
        "persons.1": "PERSON (dropped)",
    }


def test_count_unsupported_part_pii_ignores_blank_and_absent():
    plain = workbook_bytes(_workbook())  # openpyxl's app.xml has no Manager/Company, no persons
    assert count_unsupported_part_pii(plain).entity_count == 0

    blank = inject_app_props(plain, manager="", company="   ")
    assert count_unsupported_part_pii(blank).entity_count == 0


# ── reveal ───────────────────────────────────────────────


async def test_reveal_round_trip_restores_every_surface():
    from openpyxl.packaging.custom import StringProperty

    wb = _workbook()
    ws = wb["Daten"]
    wb.properties.title = "Kundenliste Anna Mueller"
    wb.custom_doc_props.append(StringProperty(name="Mandant", value="Schmidt GmbH"))
    ws["A2"].comment = Comment("Dora Klein:\nKontakt anna@example.com", "Dora Klein")
    ws.oddHeader.left.text = "Erstellt von Anna Mueller"
    data = inject_pivot(
        workbook_bytes(wb),
        refreshed_by="Dora Klein",
        fields=[("Name", ["Anna Mueller", "Ben Schulz"]), ("Betrag", None)],
        records=[[("x", 0), ("n", 10)], [("x", 1), ("n", 20)]],
    )
    originals = ["Anna Mueller", "Ben Schulz", "Dora Klein", "Schmidt GmbH", "anna@example.com"]
    detector = SubstringDetector(
        {"Anna Mueller": "PERSON", "Schmidt GmbH": "ORGANIZATION", "anna@example.com": "EMAIL"}
    )
    mapper = PseudonymMapper()

    redacted = (await pseudonymize_xlsx_smart(data, detector, mapper, language="de")).new_bytes
    assert redacted is not None
    assert_no_part_contains(redacted, originals)

    revealed = reidentify_file_bytes(redacted, _XLSX_MIME, mapper.get_mapping_summary())
    assert revealed is not None
    out = load_workbook(io.BytesIO(revealed))
    ws = out["Daten"]
    assert out.properties.creator == "Anna Mueller"
    assert out.properties.lastModifiedBy == "Dora Klein"
    assert out.properties.title == "Kundenliste Anna Mueller"
    assert out.custom_doc_props["Mandant"].value == "Schmidt GmbH"
    assert ws["A2"].value == "Anna Mueller"
    assert ws["A2"].comment.author == "Dora Klein"
    assert ws["A2"].comment.text == "Dora Klein:\nKontakt anna@example.com"
    assert ws.oddHeader.left.text == "Erstellt von Anna Mueller"
    cache = ws._pivots[0].cache
    assert _shared_values(cache) == ["Anna Mueller", "Ben Schulz"]
    assert cache.refreshedBy == "Dora Klein"


def test_reveal_placeholder_only_in_metadata_still_rewrites():
    """Regression: _reidentify_xlsx returned the input unchanged when no *cell* changed."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Notes"])
    ws.append(["leave alone"])
    wb.properties.creator = "<<PERSON_1>>"
    wb.properties.lastModifiedBy = None
    data = workbook_bytes(wb)

    revealed = reidentify_file_bytes(data, _XLSX_MIME, {"<<PERSON_1>>": "Anna Mueller"})

    assert revealed is not None
    assert revealed != data
    assert load_workbook(io.BytesIO(revealed)).properties.creator == "Anna Mueller"


# ── review round 1: pivot cache must agree with the cell pass ─


async def test_pivot_value_known_to_mapper_is_replaced_even_when_field_unclassified():
    """The sheet pass may classify a column from a row the cache sample never sees."""
    wb = _pivot_workbook(field_name="Spalte1")
    mapper = PseudonymMapper()
    anna = mapper.get_or_create("Anna Mueller", "PERSON")  # the cell pass already mapped it

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    assert _shared_values(wb["Daten"]._pivots[0].cache) == [anna, "Ben Schulz"]
    assert "pivot1.Spalte1" not in result.classifications
    # creator + lastModifiedBy + refreshedBy + the one known shared item
    assert result.entity_types == {"PERSON": 4}


async def test_pivot_field_classified_from_sheet_header_map():
    wb = _pivot_workbook(field_name="Spalte1")

    result = await pseudonymize_workbook_parts(
        wb, SubstringDetector({}), PseudonymMapper(), "de", known_fields={"spalte1": "PERSON"}
    )

    assert result.classifications["pivot1.Spalte1"] == "PERSON (sheet)"
    values = _shared_values(wb["Daten"]._pivots[0].cache)
    assert all(isinstance(v, str) and v.startswith("<<PERSON_") for v in values), values


async def test_pivots_on_different_sheets_sharing_cache_id_all_scrubbed_counted_once():
    wb = _pivot_workbook(tables=2, table_sheets=["sheet1", "sheet2"])
    ws1, ws2 = wb["Daten"], wb["Zwei"]
    # openpyxl hands each sheet its own CacheDefinition copy for the same cacheId, and its
    # writer emits every copy — so every copy must be scrubbed, but counted once.
    assert ws1._pivots[0].cache is not ws2._pivots[0].cache, "fixture: expected separate objects"
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    assert result.entity_types == {"PERSON": 5}
    assert not any(k.startswith("pivot2.") for k in result.classifications)
    ws1["A2"].value, ws1["A3"].value = "<<PERSON_1>>", "<<PERSON_3>>"
    assert_no_part_contains(workbook_bytes(wb), ["Anna Mueller", "Ben Schulz", "Dora Klein"])


async def test_pivot_captions_and_group_items_scrubbed():
    wb = _pivot_workbook(
        captions={"Anna Mueller": "Frau Mueller"}, group_items=["Gruppe1", "Ben Schulz"]
    )
    cache_field = wb["Daten"]._pivots[0].cache.cacheFields[0]
    assert cache_field.sharedItems._fields[0].c == "Frau Mueller", "fixture: caption not loaded"
    assert cache_field.fieldGroup.groupItems.s[1].v == "Ben Schulz", "fixture: group not loaded"
    mapper = PseudonymMapper()

    await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    assert cache_field.sharedItems._fields[0].c == mapper.get_or_create("Frau Mueller", "PERSON")
    assert [i.v for i in cache_field.fieldGroup.groupItems.s] == [
        mapper.get_or_create("Gruppe1", "PERSON"),
        mapper.get_or_create("Ben Schulz", "PERSON"),
    ]
    ws = wb["Daten"]
    ws["A2"].value, ws["A3"].value = "<<PERSON_1>>", "<<PERSON_3>>"
    assert_no_part_contains(
        workbook_bytes(wb), ["Anna Mueller", "Ben Schulz", "Frau Mueller", "Gruppe1"]
    )


async def test_pivot_items_differing_only_by_whitespace_stay_distinct():
    wb = _pivot_workbook(shared=["Anna Mueller", "Anna Mueller "])

    await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert _shared_values(wb["Daten"]._pivots[0].cache) == ["<<PERSON_1>>", "<<PERSON_1>> "]


# ── review round 1: authors ──────────────────────────────


async def test_comment_author_prefix_survives_partial_detector_span():
    """A detector span covering only part of the author must not leak the rest."""

    def fixture() -> Workbook:
        wb = _workbook()
        wb["Daten"]["A2"].comment = Comment("Klein, Dora Maria:\nbitte prüfen", "Klein, Dora Maria")
        return wb

    detector = SubstringDetector({"Klein, Dora": "PERSON"})
    wb = fixture()
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    author = wb["Daten"]["A2"].comment.author
    assert author.startswith("<<PERSON_")
    assert wb["Daten"]["A2"].comment.text == f"{author}:\nbitte prüfen"
    # creator + lastModifiedBy + author + prefix; the partial span is not a fifth hit
    assert result.entity_types == {"PERSON": 4}

    counted = await pseudonymize_workbook_parts(
        fixture(), detector, PseudonymMapper(), "de", apply=False
    )
    assert counted.entity_types == result.entity_types


async def test_threaded_comment_guid_author_is_not_a_person():
    """Excel's legacy mirror of a threaded comment carries author 'tc={GUID}' — not PII."""
    wb = _workbook()
    guid = "tc={5F7A1C2E-9B3D-4E6F-8A1B-2C3D4E5F6A7B}"
    wb["Daten"]["A2"].comment = Comment("[Threaded comment]\n\nComment:\n    bitte prüfen", guid)

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert wb["Daten"]["A2"].comment.author == guid
    assert result.entity_count == 2


# ── review round 1: header/footer font codes ─────────────


async def test_header_text_captured_by_font_code_is_scrubbed():
    """openpyxl's greedy &"font" parse moves text between two font codes into .font."""
    from openpyxl.worksheet.header_footer import _HeaderFooterPart

    wb = _workbook()
    ws = wb["Daten"]
    ws.oddHeader.left = _HeaderFooterPart.from_str(
        '&"-,Bold"Max Mustermann&"-,Regular" Vertraulich'
    )
    assert ws.oddHeader.left.text == " Vertraulich", "fixture: expected the greedy parse"
    detector = SubstringDetector({"Max Mustermann": "PERSON"})
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    placeholder = mapper.get_or_create("Max Mustermann", "PERSON")
    assert str(ws.oddHeader.left) == f'&"-,Bold"{placeholder}&"-,Regular" Vertraulich'
    assert result.classifications["header.odd.left@sheet1"] == "PERSON (detected)"
    ws["A2"].value = "<<PERSON_1>>"
    assert_no_part_contains(workbook_bytes(wb), ["Max Mustermann"])
