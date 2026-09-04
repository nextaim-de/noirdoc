"""Fast characterization tests for ``pseudonymize_xlsx_smart`` (no ML models).

The SDK-level XLSX tests in ``tests/test_sdk_xlsx.py`` are ``slow`` and excluded from
CI, so the three-tier column pipeline had no fast coverage. These pin its behaviour
so the part-level work in ``xlsx_parts`` can refactor Tier 2 safely.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from noirdoc.file_analysis.xlsx_inference import XlsxLoadError, pseudonymize_xlsx_smart
from noirdoc.file_reidentification.service import reidentify_file_bytes
from noirdoc.pseudonymization.mapper import PseudonymMapper
from tests.file_analysis.xlsx_helpers import (
    SubstringDetector,
    assert_no_part_contains,
    inject_app_props,
    inject_pivot,
    inject_threaded_comment_parts,
    part_names,
    read_part,
    rewrite_zip,
    strip_core_creator,
    workbook_bytes,
)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _sheet_bytes(
    header: list[str], rows: list[list[object]], *, creator: str | None = None
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(header)
    for row in rows:
        ws.append(row)
    # openpyxl defaults creator to "openpyxl" — a PERSON hit by design (no allowlist) — and
    # its *reader* substitutes that default again when the element is absent. An empty
    # element round-trips as None, so that is how a fixture gets "no creator".
    wb.properties.creator = creator if creator is not None else ""
    return workbook_bytes(wb)


async def test_header_tier1_pseudonymizes_classified_columns():
    data = _sheet_bytes(
        ["Name", "Email", "Notes"],
        [
            ["Anna Mueller", "anna@example.com", "leave alone"],
            ["Ben Schulz", "ben@example.com", "x"],
        ],
    )
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, SubstringDetector({}), mapper, language="de")

    assert result.new_bytes is not None
    assert result.entity_count == 4
    assert result.entity_types == {"PERSON": 2, "EMAIL": 2}
    assert result.column_classifications == {"Name": "PERSON (header)", "Email": "EMAIL (header)"}
    ws = load_workbook(io.BytesIO(result.new_bytes))["Daten"]
    assert ws.cell(2, 1).value.startswith("<<PERSON_")
    assert ws.cell(3, 1).value.startswith("<<PERSON_")
    assert ws.cell(2, 2).value.startswith("<<EMAIL_")
    assert ws.cell(3, 2).value.startswith("<<EMAIL_")
    assert ws.cell(2, 3).value == "leave alone"
    assert ws.cell(1, 1).value == "Name"


async def test_sample_tier2_classifies_unlabeled_column():
    names = [
        "Anna Mueller",
        "Ben Schulz",
        "Cara Weiss",
        "Dora Klein",
        "Emil Roth",
        "Finn Lang",
        "Gerd Haas",
    ]
    data = _sheet_bytes(["Spalte A"], [[n] for n in names])
    # Only the first name is "detectable" — it sits inside the 5-row sample window, which
    # must classify the whole column so rows beyond the sample get replaced too.
    detector = SubstringDetector({"Anna Mueller": "PERSON"})
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, detector, mapper, language="de")

    assert result.column_classifications == {"Spalte A": "PERSON (sampled)"}
    assert result.entity_count == len(names)
    assert result.new_bytes is not None
    ws = load_workbook(io.BytesIO(result.new_bytes))["Daten"]
    for row in range(2, 2 + len(names)):
        assert ws.cell(row, 1).value.startswith("<<PERSON_"), row


async def test_count_only_mode_does_not_touch_mapper():
    data = _sheet_bytes(["Name"], [["Anna Mueller"], ["Ben Schulz"]])
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(
        data, SubstringDetector({}), mapper, language="de", pseudonymize=False
    )

    assert result.new_bytes is None
    assert result.entity_count == 2
    assert mapper.entity_count == 0


async def test_no_pii_returns_no_bytes():
    data = _sheet_bytes(["Notes"], [["leave alone"], ["also untouched"]])
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, SubstringDetector({}), mapper, language="de")

    assert result.new_bytes is None
    assert result.entity_count == 0
    assert mapper.entity_count == 0


# ── parts outside the cell grid (xlsx_parts wiring) ──────


async def test_metadata_only_workbook_is_rewritten():
    """Regression: a workbook whose only PII is metadata used to pass through byte-identical."""
    data = _sheet_bytes(["Notes"], [["leave alone"]], creator="Anna Mueller")
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, SubstringDetector({}), mapper, language="de")

    assert result.new_bytes is not None
    assert result.entity_count == 1
    assert result.entity_types == {"PERSON": 1}
    assert result.column_classifications["core.creator"] == "PERSON (forced)"
    assert load_workbook(io.BytesIO(result.new_bytes)).properties.creator == "<<PERSON_1>>"
    assert_no_part_contains(result.new_bytes, ["Anna Mueller"])


async def test_count_only_mode_counts_metadata_without_mutating():
    data = _sheet_bytes(["Notes"], [["leave alone"]], creator="Anna Mueller")
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(
        data, SubstringDetector({}), mapper, language="de", pseudonymize=False
    )

    assert result.new_bytes is None
    assert result.entity_count == 1
    assert mapper.entity_count == 0


async def test_unsupported_parts_are_dropped_and_counted():
    """app.xml Manager/Company and xl/persons are dropped by the writer — but only if a
    save happens, so they must count as hits even though nothing gets mapped."""
    data = _sheet_bytes(["Notes"], [["leave alone"]])
    data = inject_app_props(data, manager="Dora Klein", company="Schmidt GmbH")
    data = inject_threaded_comment_parts(data, display_name="Emil Roth", text="bitte prüfen")
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, SubstringDetector({}), mapper, language="de")

    assert result.new_bytes is not None
    names = part_names(result.new_bytes)
    assert not any(n.startswith(("xl/threadedComments/", "xl/persons/")) for n in names)
    assert_no_part_contains(result.new_bytes, ["Dora Klein", "Schmidt GmbH", "Emil Roth"])
    assert result.entity_types == {"PERSON": 2, "ORGANIZATION": 1}
    assert result.column_classifications["app.Manager"] == "PERSON (dropped)"
    assert result.column_classifications["app.Company"] == "ORGANIZATION (dropped)"
    assert result.column_classifications["persons.1"] == "PERSON (dropped)"  # never the name
    assert mapper.entity_count == 0  # dropped, not mapped — nothing to reveal


async def test_pivot_cache_matches_pseudonymized_cells():
    """The leger leak: sheet cells got placeholders while the pivot cache kept the originals."""
    data = _sheet_bytes(["Name", "Betrag"], [["Anna Mueller", 10], ["Ben Schulz", 20]])
    data = inject_pivot(
        data,
        refreshed_by="Dora Klein",
        fields=[("Name", ["Anna Mueller", "Ben Schulz"]), ("Betrag", None)],
        records=[[("x", 0), ("n", 10)], [("x", 1), ("n", 20)]],
    )
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, SubstringDetector({}), mapper, language="de")

    assert result.new_bytes is not None
    assert_no_part_contains(result.new_bytes, ["Anna Mueller", "Ben Schulz", "Dora Klein"])
    ws = load_workbook(io.BytesIO(result.new_bytes))["Daten"]
    cache = ws._pivots[0].cache
    shared = [item.v for item in cache.cacheFields[0].sharedItems._fields]
    assert shared == [ws["A2"].value, ws["A3"].value]
    assert result.entity_types == {"PERSON": 5}  # 2 cells + refreshedBy + 2 shared items


# ── fail closed on load failures (issue #13) ─────────────


async def test_pivot_discrete_pr_load_failure_raises_instead_of_failing_open():
    """openpyxl 3.1.5 crashes on ``cacheField/fieldGroup/discretePr`` (manually grouped
    pivot fields). The old code swallowed that and returned an empty result, which the
    callers turned into 'the original workbook is the redacted output'."""
    data = _sheet_bytes(["Name", "Betrag"], [["Anna Mueller", 10], ["Ben Schulz", 20]])
    data = inject_pivot(
        data,
        refreshed_by="Dora Klein",
        fields=[("Name", ["Anna Mueller", "Ben Schulz"]), ("Betrag", None)],
        records=[[("x", 0), ("n", 10)], [("x", 1), ("n", 20)]],
        group_items=["Gruppe A", "Gruppe B"],
    )
    part = "xl/pivotCache/pivotCacheDefinition1.xml"
    xml = read_part(data, part)
    assert b"<groupItems" in xml, "fixture: pivot cache has no fieldGroup to poison"
    xml = xml.replace(
        b"<groupItems",
        b'<discretePr count="2"><x v="0"/><x v="1"/></discretePr><groupItems',
        1,
    )
    data = rewrite_zip(data, {part: xml})
    mapper = PseudonymMapper()

    with pytest.raises(XlsxLoadError, match="cannot load xlsx workbook"):
        await pseudonymize_xlsx_smart(data, SubstringDetector({}), mapper, language="de")

    assert mapper.entity_count == 0


# ── review round 1 ───────────────────────────────────────


async def test_pivot_cache_scrubbed_when_only_the_sheet_sample_hits():
    """Column classified from row 2, but the cache lists that value beyond its sample window."""
    names = ["Anna Mueller", "Ben Schulz", "Cara Weiss", "Dora Klein", "Emil Roth", "Finn Lang"]
    data = _sheet_bytes(["Spalte1", "Betrag"], [[n, i] for i, n in enumerate(names)])
    data = inject_pivot(
        data,
        refreshed_by="Gerd Haas",
        fields=[("Spalte1", list(reversed(names))), ("Betrag", None)],
        records=[[("x", i), ("n", i)] for i in range(len(names))],
        source_ref="A1:B7",
    )
    detector = SubstringDetector({"Anna Mueller": "PERSON"})

    result = await pseudonymize_xlsx_smart(data, detector, PseudonymMapper(), language="de")

    assert result.column_classifications["Spalte1"] == "PERSON (sampled)"
    assert result.column_classifications["pivot1.Spalte1"] == "PERSON (sheet)"
    assert result.new_bytes is not None
    assert_no_part_contains(result.new_bytes, [*names, "Gerd Haas"])


async def test_absent_creator_element_is_not_a_phantom_person():
    """openpyxl's reader substitutes creator='openpyxl' when the element is missing."""
    data = strip_core_creator(_sheet_bytes(["Notes"], [["leave alone"]], creator="Anna Mueller"))
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, SubstringDetector({}), mapper, language="de")

    assert result.entity_count == 0
    assert result.new_bytes is None
    assert mapper.entity_count == 0


async def test_chartsheet_is_skipped_by_the_cell_pass_and_its_header_scrubbed():
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Name", "Betrag"])
    ws.append(["Anna Mueller", 10])
    ws.append(["Ben Schulz", 20])
    wb.properties.creator = ""
    chartsheet = wb.create_chartsheet("Chart")
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    chartsheet.add_chart(chart)
    chartsheet.oddHeader.left.text = "Erstellt von Anna Mueller"
    data = workbook_bytes(wb)
    assert "Chart" in load_workbook(io.BytesIO(data)).sheetnames, "fixture: chartsheet lost"
    detector = SubstringDetector({"Anna Mueller": "PERSON"})
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(data, detector, mapper, language="de")

    assert result.new_bytes is not None
    out = load_workbook(io.BytesIO(result.new_bytes))
    assert out["Daten"]["A2"].value.startswith("<<PERSON_")
    assert out["Chart"].oddHeader.left.text == "Erstellt von <<PERSON_1>>"

    revealed = reidentify_file_bytes(result.new_bytes, _XLSX_MIME, mapper.get_mapping_summary())
    assert revealed is not None
    back = load_workbook(io.BytesIO(revealed))
    assert back["Daten"]["A2"].value == "Anna Mueller"
    assert back["Chart"].oddHeader.left.text == "Erstellt von Anna Mueller"


# ── issue #11: bound the free-text NER cost in detect/block modes ──


async def test_block_mode_cell_hit_skips_part_free_text_scan():
    """A cell-grid hit settles the block decision — comment NER is skipped but reported."""
    from openpyxl.comments import Comment

    from noirdoc.detection.base import DetectedEntity

    calls: list[str] = []

    class CountingDetector(SubstringDetector):
        async def detect(self, text: str, language: str = "de") -> list[DetectedEntity]:
            calls.append(text)
            return await super().detect(text, language)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Name", "Betrag"])
    ws.append(["Anna Mueller", 10])
    # A threaded-mirror author is excluded from author handling — the only other
    # entity is the cell hit, so the skip below must come from the cell short-circuit.
    ws["B2"].comment = Comment(
        "Geheimnotiz Ben Schulz", "tc={5F7A1C2E-9B3D-4E6F-8A1B-2C3D4E5F6A7B}"
    )
    wb.properties.creator = ""  # absent creator: no deterministic author hit
    data = workbook_bytes(wb)

    result = await pseudonymize_xlsx_smart(
        data,
        CountingDetector({"Ben Schulz": "PERSON"}),
        PseudonymMapper(),
        language="de",
        pseudonymize=False,
        stop_on_first_hit=True,
    )

    assert calls == []  # the Name column is classified by header keyword, no NER at all
    assert result.entity_count == 1
    assert result.free_texts_skipped == 1
    assert result.new_bytes is None


async def test_detect_mode_cap_is_surfaced_on_result():
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Betrag"])
    ws.append([10])
    for row in (2, 3, 4):
        ws.cell(row=row, column=2).comment = Comment(f"Kommentar {row}", "")
    wb.properties.creator = ""
    data = workbook_bytes(wb)

    result = await pseudonymize_xlsx_smart(
        data,
        SubstringDetector({}),
        PseudonymMapper(),
        language="de",
        pseudonymize=False,
        max_free_texts=2,
    )

    assert result.free_texts_skipped == 1


# ── row 1: header label or first record? (issue #30) ─────


def _rows_bytes(rows: list[list[object]], *, title: str = "Daten") -> bytes:
    """A sheet built from *rows* with no header row prepended."""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    wb.properties.creator = ""
    return workbook_bytes(wb)


async def test_row_one_is_redacted_when_the_column_below_says_it_is_data():
    """A list pasted without a header shipped its whole first record."""
    data = _rows_bytes([["Anna Mueller"], ["Ben Schulz"], ["Carla Weber"]])
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(
        data,
        SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"}),
        mapper,
        language="de",
    )

    assert result.new_bytes is not None
    assert_no_part_contains(result.new_bytes, ["Anna Mueller"])
    ws = load_workbook(io.BytesIO(result.new_bytes))["Daten"]
    assert ws["A1"].value == "<<PERSON_1>>"
    assert ws["A2"].value == "<<PERSON_2>>"
    # Row 1 and row 5 share the mapper, so a repeat gets the same placeholder.
    assert mapper.reverse_lookup("<<PERSON_1>>") == "Anna Mueller"


async def test_single_row_sheet_is_no_longer_skipped():
    """``max_row < 2`` used to skip the sheet before classification even ran."""
    data = _rows_bytes([["Frida Sologne"]], title="Einzeilig")
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(
        data, SubstringDetector({"Frida Sologne": "PERSON"}), mapper, language="de"
    )

    assert result.new_bytes is not None
    assert_no_part_contains(result.new_bytes, ["Frida Sologne"])
    assert load_workbook(io.BytesIO(result.new_bytes))["Einzeilig"]["A1"].value == "<<PERSON_1>>"


async def test_a_recognized_header_row_is_left_alone():
    """The control: over-redacting headers would be its own kind of damage."""
    data = _sheet_bytes(
        ["Name", "Ansprechpartner"],
        [["Anna Mueller", "Ben Schulz"], ["Carla Weber", "Dora Klein"]],
    )

    result = await pseudonymize_xlsx_smart(
        data,
        # A detector that would happily flag the header words too.
        SubstringDetector({"Name": "PERSON", "Ansprechpartner": "PERSON"}),
        PseudonymMapper(),
        language="de",
    )

    assert result.new_bytes is not None
    ws = load_workbook(io.BytesIO(result.new_bytes))["Daten"]
    assert ws["A1"].value == "Name"
    assert ws["B1"].value == "Ansprechpartner"


async def test_an_unrecognized_header_over_unclassified_data_is_left_alone():
    """ "Betrag" is not in the keyword map and sits over numbers.

    The column never gets classified, so row 1 is not a candidate — a stray NER
    hit on a lone header word cannot destroy it.
    """
    data = _sheet_bytes(["Betrag"], [[10], [20]])

    result = await pseudonymize_xlsx_smart(
        data, SubstringDetector({"Betrag": "PERSON"}), PseudonymMapper(), language="de"
    )

    assert result.new_bytes is None  # nothing changed at all
    assert load_workbook(io.BytesIO(data))["Daten"]["A1"].value == "Betrag"


async def test_row_one_data_never_becomes_a_classification_label():
    """``column_classifications`` is logged, so a data value must not be a key."""
    data = _rows_bytes([["Anna Mueller"], ["Ben Schulz"]])

    result = await pseudonymize_xlsx_smart(
        data,
        SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"}),
        PseudonymMapper(),
        language="de",
    )

    labels = " ".join(result.column_classifications)
    assert "Anna Mueller" not in labels
    assert "col1" in result.column_classifications
    assert result.column_classifications["row1!col1"] == "PERSON (row 1 data)"


async def test_row_one_round_trips_through_reveal():
    data = _rows_bytes([["Anna Mueller"], ["Ben Schulz"]])
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(
        data,
        SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"}),
        mapper,
        language="de",
    )
    assert result.new_bytes is not None

    revealed = reidentify_file_bytes(result.new_bytes, _XLSX_MIME, mapper.get_mapping_summary())
    assert revealed is not None
    ws = load_workbook(io.BytesIO(revealed))["Daten"]
    assert ws["A1"].value == "Anna Mueller"
    assert ws["A2"].value == "Ben Schulz"


async def test_detect_mode_counts_row_one_without_writing():
    data = _rows_bytes([["Anna Mueller"], ["Ben Schulz"]])
    mapper = PseudonymMapper()

    result = await pseudonymize_xlsx_smart(
        data,
        SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"}),
        mapper,
        language="de",
        pseudonymize=False,
    )

    assert result.new_bytes is None
    assert result.entity_types == {"PERSON": 2}  # row 1 counted, not just row 2
    assert mapper.get_mapping_summary() == {}


async def test_unrecognized_header_over_classified_data_survives_unless_flagged():
    """The residual case, pinned deliberately.

    "Zustaendig" is not in the keyword map but sits over names, so the column
    *is* classified and row 1 becomes a candidate. It is rewritten only if the
    detector flags the header word itself — which is the safe direction, and
    the reason the candidate rule leans on the column's classification rather
    than redacting row 1 outright.
    """
    data = _sheet_bytes(["Zustaendig"], [["Anna Mueller"], ["Ben Schulz"]])

    result = await pseudonymize_xlsx_smart(
        data,
        SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"}),
        PseudonymMapper(),
        language="de",
    )

    assert result.new_bytes is not None
    ws = load_workbook(io.BytesIO(result.new_bytes))["Daten"]
    assert ws["A1"].value == "Zustaendig"  # no hit on the word, header intact
    assert ws["A2"].value == "<<PERSON_1>>"
    # ... and the header still names the column for the pivot pass.
    assert result.column_classifications["Zustaendig"] == "PERSON (sampled)"


async def test_merged_row_one_does_not_crash():
    """Merged title rows are common; MergedCell.value is read-only."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws["A1"] = "Akte Anna Mueller"
    ws.merge_cells("A1:C1")
    ws.append(["Ben Schulz", "x", "y"])
    ws.append(["Carla Weber", "x", "y"])
    wb.properties.creator = ""
    data = workbook_bytes(wb)

    result = await pseudonymize_xlsx_smart(
        data,
        SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"}),
        PseudonymMapper(),
        language="de",
    )

    assert result.new_bytes is not None
    ws_out = load_workbook(io.BytesIO(result.new_bytes))["Daten"]
    # Only the anchor cell of a merged range holds the value; the rest read None
    # and are filtered out before any write is attempted.
    assert ws_out["A1"].value == "Akte <<PERSON_1>>"
    assert_no_part_contains(result.new_bytes, ["Anna Mueller"])
