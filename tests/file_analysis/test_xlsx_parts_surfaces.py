"""Issue #15: the remaining verbatim-round-tripped XLSX surfaces.

Covers chart string caches and titles, hyperlink tooltip/display and
``mailto:`` targets, AutoFilter criteria, conditional-formatting text
literals, data-validation inline lists and prompts, and pivot-table
captions / label filters — each as slot kinds in ``iter_text_slots`` so
redact and reveal stay symmetric.
"""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import CustomFilter, CustomFilters, FilterColumn, Filters
from openpyxl.worksheet.hyperlink import Hyperlink

from noirdoc.file_analysis.xlsx_inference import pseudonymize_xlsx_smart
from noirdoc.file_analysis.xlsx_parts import pseudonymize_workbook_parts
from noirdoc.file_reidentification.service import reidentify_file_bytes
from noirdoc.pseudonymization.mapper import PseudonymMapper
from tests.file_analysis.xlsx_helpers import (
    SubstringDetector,
    assert_no_part_contains,
    inject_chart,
    inject_pivot,
    read_part,
    workbook_bytes,
)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Name", "Betrag"])
    ws.append(["Anna Mueller", 10])
    ws.append(["Ben Schulz", 20])
    wb.properties.creator = None
    wb.properties.lastModifiedBy = None
    return wb


# ── chart string caches and titles ───────────────────────


def _chart_workbook() -> bytes:
    return inject_chart(
        workbook_bytes(_workbook()),
        title="Umsatz Anna Mueller",
        series_name="Betrag Anna Mueller",
        categories=["Anna Mueller", "Ben Schulz"],
        axis_title="Kunde Ben Schulz",
    )


async def test_chart_caches_and_titles_scrubbed_and_survive_save():
    wb = load_workbook(io.BytesIO(_chart_workbook()))
    detector = SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"})
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    anna = mapper.get_or_create("Anna Mueller", "PERSON")
    ben = mapper.get_or_create("Ben Schulz", "PERSON")
    chart = wb["Daten"]._charts[0]
    assert chart.title.tx.rich.p[0].r[0].t == f"Umsatz {anna}"
    assert chart.x_axis.title.tx.rich.p[0].r[0].t == f"Kunde {ben}"
    series = chart.series[0]
    assert series.tx.strRef.strCache.pt[0].v == f"Betrag {anna}"
    assert [pt.v for pt in series.cat.strRef.strCache.pt] == [anna, ben]
    assert [pt.v for pt in series.val.numRef.numCache.pt] == [10.0, 20.0]  # numbers untouched
    assert result.classifications["chart1.title@sheet1"] == "PERSON (detected)"

    # openpyxl re-serializes chart parts it understands — the rewrite must survive that.
    ws = wb["Daten"]
    ws["A2"].value, ws["A3"].value = anna, ben
    out = workbook_bytes(wb)
    assert_no_part_contains(out, ["Anna Mueller", "Ben Schulz"])
    reloaded = load_workbook(io.BytesIO(out))["Daten"]._charts[0]
    assert [pt.v for pt in reloaded.series[0].cat.strRef.strCache.pt] == [anna, ben]


async def test_chart_cache_value_known_to_mapper_reuses_cell_placeholder():
    wb = load_workbook(io.BytesIO(_chart_workbook()))
    mapper = PseudonymMapper()
    anna = mapper.get_or_create("Anna Mueller", "PERSON")  # the cell pass already mapped it

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), mapper, "de")

    series = wb["Daten"]._charts[0].series[0]
    assert [pt.v for pt in series.cat.strRef.strCache.pt] == [anna, "Ben Schulz"]
    assert result.classifications["chart1.series1.cat@sheet1"] == "PERSON (mapped)"
    # title/series name are free text: no detector hit, no mapper full-match — left alone.
    assert wb["Daten"]._charts[0].title.tx.rich.p[0].r[0].t == "Umsatz Anna Mueller"


# ── hyperlinks ───────────────────────────────────────────


def _hyperlink_workbook() -> Workbook:
    wb = _workbook()
    wb["Daten"]["A2"].hyperlink = Hyperlink(
        ref="A2",
        target="mailto:anna@example.com",
        tooltip="Mail an Anna Mueller",
        display="Anna Mueller",
    )
    return wb


async def test_hyperlink_tooltip_display_and_mailto_target_scrubbed():
    wb = _hyperlink_workbook()
    detector = SubstringDetector({"Anna Mueller": "PERSON"})
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    link = wb["Daten"]["A2"].hyperlink
    person = mapper.get_or_create("Anna Mueller", "PERSON")
    assert link.tooltip == f"Mail an {person}"
    assert link.display == person
    # <</>>  are illegal URI characters: the placeholder is percent-encoded instead.
    assert link.target == "mailto:%3C%3CEMAIL_1%3E%3E"
    assert mapper.reverse_lookup("<<EMAIL_1>>") == "anna@example.com"
    assert result.classifications["hyperlink.mailto@sheet1!A2"] == "EMAIL (forced)"

    wb["Daten"]["A2"].value = person
    out = workbook_bytes(wb)
    assert_no_part_contains(out, ["Anna Mueller", "anna@example.com"])
    rels = read_part(out, "xl/worksheets/_rels/sheet1.xml.rels").decode()
    assert "mailto:%3C%3CEMAIL_1%3E%3E" in rels


async def test_hyperlink_mailto_reveal_restores_plain_target():
    wb = _hyperlink_workbook()
    mapper = PseudonymMapper()
    await pseudonymize_workbook_parts(
        wb, SubstringDetector({"Anna Mueller": "PERSON"}), mapper, "de"
    )
    wb["Daten"]["A2"].value = mapper.get_or_create("Anna Mueller", "PERSON")
    redacted = workbook_bytes(wb)

    revealed = reidentify_file_bytes(redacted, _XLSX_MIME, mapper.get_mapping_summary())

    assert revealed is not None
    out = load_workbook(io.BytesIO(revealed))["Daten"]
    link = out["A2"].hyperlink
    assert link.target == "mailto:anna@example.com"
    assert link.tooltip == "Mail an Anna Mueller"
    assert link.display == "Anna Mueller"


async def test_non_mailto_hyperlink_target_left_alone():
    wb = _workbook()
    wb["Daten"]["A2"].hyperlink = Hyperlink(ref="A2", target="https://example.com/anna")

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert wb["Daten"]["A2"].hyperlink.target == "https://example.com/anna"
    assert result.entity_count == 0


# ── autofilter criteria ──────────────────────────────────


async def test_autofilter_criteria_scrubbed():
    wb = _workbook()
    ws = wb["Daten"]
    ws.auto_filter.ref = "A1:B3"
    ws.auto_filter.filterColumn.append(
        FilterColumn(colId=0, filters=Filters(filter=["Anna Mueller", "unbekannt"]))
    )
    ws.auto_filter.filterColumn.append(
        FilterColumn(
            colId=1,
            customFilters=CustomFilters(
                customFilter=[CustomFilter(operator="equal", val="*Ben Schulz*")]
            ),
        )
    )
    detector = SubstringDetector({"Ben Schulz": "PERSON"})
    mapper = PseudonymMapper()
    anna = mapper.get_or_create("Anna Mueller", "PERSON")  # cell pass knows the exact value

    result = await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    ben = mapper.get_or_create("Ben Schulz", "PERSON")
    columns = ws.auto_filter.filterColumn
    assert list(columns[0].filters.filter) == [anna, "unbekannt"]  # unknown value stays
    assert columns[1].customFilters.customFilter[0].val == f"*{ben}*"  # wildcards survive
    assert result.entity_types == {"PERSON": 2}

    ws["A2"].value, ws["A3"].value = anna, ben
    assert_no_part_contains(workbook_bytes(wb), ["Anna Mueller", "Ben Schulz"])


# ── conditional formatting ───────────────────────────────


async def test_conditional_formatting_text_and_formula_kept_in_lockstep():
    wb = _workbook()
    ws = wb["Daten"]
    dxf = DifferentialStyle(font=Font(bold=True))
    ws.conditional_formatting.add(
        "A1:A10",
        Rule(
            type="containsText",
            operator="containsText",
            text="Anna Mueller",
            formula=['NOT(ISERROR(SEARCH("Anna Mueller",A1)))'],
            dxf=dxf,
        ),
    )
    ws.conditional_formatting.add(
        "A1:A10",
        Rule(type="cellIs", operator="equal", formula=['"Ben Schulz"'], dxf=dxf),
    )
    detector = SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"})
    mapper = PseudonymMapper()

    await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    anna = mapper.get_or_create("Anna Mueller", "PERSON")
    ben = mapper.get_or_create("Ben Schulz", "PERSON")
    rules = [rule for fmt in ws.conditional_formatting for rule in fmt.rules]
    assert rules[0].text == anna
    assert rules[0].formula == [f'NOT(ISERROR(SEARCH("{anna}",A1)))']
    assert rules[1].formula == [f'"{ben}"']

    ws["A2"].value, ws["A3"].value = anna, ben
    redacted = workbook_bytes(wb)
    assert_no_part_contains(redacted, ["Anna Mueller", "Ben Schulz"])

    revealed = reidentify_file_bytes(redacted, _XLSX_MIME, mapper.get_mapping_summary())
    assert revealed is not None
    out = load_workbook(io.BytesIO(revealed))["Daten"]
    rules = [rule for fmt in out.conditional_formatting for rule in fmt.rules]
    assert rules[0].text == "Anna Mueller"
    assert rules[0].formula == ['NOT(ISERROR(SEARCH("Anna Mueller",A1)))']
    assert rules[1].formula == ['"Ben Schulz"']


# ── data validation ──────────────────────────────────────


async def test_data_validation_inline_list_and_prompts_scrubbed():
    wb = _workbook()
    ws = wb["Daten"]
    dv = DataValidation(type="list", formula1='"Anna Mueller,Ben Schulz,offen"', allow_blank=True)
    dv.promptTitle = "Anna Mueller"
    dv.prompt = "Bitte Anna Mueller waehlen"
    dv.errorTitle = "Fehler"
    dv.error = "Nur Ben Schulz ist erlaubt"
    dv.add("C2:C10")
    ws.add_data_validation(dv)
    detector = SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"})
    mapper = PseudonymMapper()

    await pseudonymize_workbook_parts(wb, detector, mapper, "de")

    anna = mapper.get_or_create("Anna Mueller", "PERSON")
    ben = mapper.get_or_create("Ben Schulz", "PERSON")
    assert dv.formula1 == f'"{anna},{ben},offen"'
    assert dv.promptTitle == anna
    assert dv.prompt == f"Bitte {anna} waehlen"
    assert dv.errorTitle == "Fehler"
    assert dv.error == f"Nur {ben} ist erlaubt"

    ws["A2"].value, ws["A3"].value = anna, ben
    redacted = workbook_bytes(wb)
    assert_no_part_contains(redacted, ["Anna Mueller", "Ben Schulz"])

    revealed = reidentify_file_bytes(redacted, _XLSX_MIME, mapper.get_mapping_summary())
    assert revealed is not None
    out = load_workbook(io.BytesIO(revealed))["Daten"]
    dv_out = out.data_validations.dataValidation[0]
    assert dv_out.formula1 == '"Anna Mueller,Ben Schulz,offen"'
    assert dv_out.prompt == "Bitte Anna Mueller waehlen"


async def test_data_validation_range_reference_left_alone():
    wb = _workbook()
    ws = wb["Daten"]
    dv = DataValidation(type="list", formula1="$A$2:$A$3")
    dv.add("C2:C10")
    ws.add_data_validation(dv)

    result = await pseudonymize_workbook_parts(wb, SubstringDetector({}), PseudonymMapper(), "de")

    assert dv.formula1 == "$A$2:$A$3"
    assert result.entity_count == 0


# ── pivot-table captions and label filters ───────────────


async def test_pivot_table_captions_and_label_filters_scrubbed():
    wb = _workbook()
    xlsx = inject_pivot(
        workbook_bytes(wb),
        refreshed_by="Dora Klein",
        fields=[("Name", ["Anna Mueller", "Ben Schulz"]), ("Betrag", None)],
        records=[[("x", 0), ("n", 10)], [("x", 1), ("n", 20)]],
        data_caption="Werte Anna Mueller",
        item_captions={0: "Frau Anna Mueller"},
        page_field=("Anna Mueller", "Nur Anna Mueller"),
        label_filter="Ben Schulz",
    )
    loaded = load_workbook(io.BytesIO(xlsx))
    pivot = loaded["Daten"]._pivots[0]
    assert pivot.filters[0].stringValue1 == "Ben Schulz", "fixture: filter not loaded"
    detector = SubstringDetector({"Anna Mueller": "PERSON"})
    mapper = PseudonymMapper()

    result = await pseudonymize_workbook_parts(loaded, detector, mapper, "de")

    anna = mapper.get_or_create("Anna Mueller", "PERSON")
    ben = mapper.get_or_create("Ben Schulz", "PERSON")
    frau = mapper.get_or_create("Frau Anna Mueller", "PERSON")
    nur = mapper.get_or_create("Nur Anna Mueller", "PERSON")
    assert pivot.dataCaption == f"Werte {anna}"  # free text: offset substitution
    assert pivot.pivotFields[0].items[0].n == frau  # classified field: whole caption
    assert pivot.pageFields[0].name == anna
    assert pivot.pageFields[0].cap == nur
    pivot_filter = pivot.filters[0]
    assert pivot_filter.stringValue1 == ben
    nested = pivot_filter.autoFilter.filterColumn[0].customFilters.customFilter[0]
    assert nested.val == ben  # Excel's mirror of stringValue1
    assert result.classifications["pivot1.Name"] == "PERSON (header)"

    ws = loaded["Daten"]
    ws["A2"].value, ws["A3"].value = anna, ben
    redacted = workbook_bytes(loaded)
    assert_no_part_contains(redacted, ["Anna Mueller", "Ben Schulz", "Dora Klein", "Frau", "Nur"])

    revealed = reidentify_file_bytes(redacted, _XLSX_MIME, mapper.get_mapping_summary())
    assert revealed is not None
    out = load_workbook(io.BytesIO(revealed))["Daten"]._pivots[0]
    assert out.dataCaption == "Werte Anna Mueller"
    assert out.pivotFields[0].items[0].n == "Frau Anna Mueller"
    assert out.pageFields[0].cap == "Nur Anna Mueller"
    assert out.filters[0].stringValue1 == "Ben Schulz"
    assert out.filters[0].autoFilter.filterColumn[0].customFilters.customFilter[0].val == (
        "Ben Schulz"
    )


async def test_pivot_table_captions_on_unclassified_field_left_alone():
    wb = _workbook()
    ws = wb["Daten"]
    ws["A1"] = "Notiz"
    xlsx = inject_pivot(
        workbook_bytes(wb),
        refreshed_by="Dora Klein",
        fields=[("Notiz", ["frei", "auch frei"]), ("Betrag", None)],
        records=[[("x", 0), ("n", 10)], [("x", 1), ("n", 20)]],
        item_captions={0: "ganz frei"},
        label_filter="frei",
    )
    loaded = load_workbook(io.BytesIO(xlsx))

    await pseudonymize_workbook_parts(loaded, SubstringDetector({}), PseudonymMapper(), "de")

    pivot = loaded["Daten"]._pivots[0]
    assert pivot.pivotFields[0].items[0].n == "ganz frei"
    assert pivot.filters[0].stringValue1 == "frei"


# ── full pipeline round-trip over every new surface ──────


async def test_smart_pipeline_round_trip_covers_new_surfaces():
    wb = _workbook()
    ws = wb["Daten"]
    ws["A2"].hyperlink = Hyperlink(
        ref="A2", target="mailto:anna@example.com", tooltip="Mail an Anna Mueller"
    )
    ws.auto_filter.ref = "A1:B3"
    ws.auto_filter.filterColumn.append(
        FilterColumn(colId=0, filters=Filters(filter=["Anna Mueller"]))
    )
    dv = DataValidation(type="list", formula1='"Anna Mueller,Ben Schulz"')
    dv.add("C2:C10")
    ws.add_data_validation(dv)
    ws.conditional_formatting.add(
        "A1:A10",
        Rule(
            type="containsText",
            operator="containsText",
            text="Ben Schulz",
            formula=['NOT(ISERROR(SEARCH("Ben Schulz",A1)))'],
            dxf=DifferentialStyle(font=Font(bold=True)),
        ),
    )
    data = inject_chart(
        workbook_bytes(wb),
        title="Umsatz Anna Mueller",
        series_name="Betrag",
        categories=["Anna Mueller", "Ben Schulz"],
    )
    data = inject_pivot(
        data,
        refreshed_by="Dora Klein",
        fields=[("Name", ["Anna Mueller", "Ben Schulz"]), ("Betrag", None)],
        records=[[("x", 0), ("n", 10)], [("x", 1), ("n", 20)]],
        label_filter="Anna Mueller",
    )
    originals = ["Anna Mueller", "Ben Schulz", "Dora Klein", "anna@example.com"]
    detector = SubstringDetector({"Anna Mueller": "PERSON", "Ben Schulz": "PERSON"})
    mapper = PseudonymMapper()

    redacted = (await pseudonymize_xlsx_smart(data, detector, mapper, language="de")).new_bytes
    assert redacted is not None
    assert_no_part_contains(redacted, originals)

    revealed = reidentify_file_bytes(redacted, _XLSX_MIME, mapper.get_mapping_summary())
    assert revealed is not None
    out = load_workbook(io.BytesIO(revealed))
    ws_out = out["Daten"]
    assert ws_out["A2"].value == "Anna Mueller"
    assert ws_out["A2"].hyperlink.target == "mailto:anna@example.com"
    assert ws_out["A2"].hyperlink.tooltip == "Mail an Anna Mueller"
    assert list(ws_out.auto_filter.filterColumn[0].filters.filter) == ["Anna Mueller"]
    assert ws_out.data_validations.dataValidation[0].formula1 == '"Anna Mueller,Ben Schulz"'
    rules = [rule for fmt in ws_out.conditional_formatting for rule in fmt.rules]
    assert rules[0].text == "Ben Schulz"
    chart = ws_out._charts[0]
    assert chart.title.tx.rich.p[0].r[0].t == "Umsatz Anna Mueller"
    assert [pt.v for pt in chart.series[0].cat.strRef.strCache.pt] == ["Anna Mueller", "Ben Schulz"]
    assert ws_out._pivots[0].filters[0].stringValue1 == "Anna Mueller"
