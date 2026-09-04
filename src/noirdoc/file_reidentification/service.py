"""Reidentify pseudonyms in downloaded files.

Replaces <<TYPE_N>> tokens with original values in supported formats:

* **DOCX** – every text surface via the shared walker in
  :mod:`noirdoc.file_analysis.docx_surfaces` (body incl. content controls,
  nested tables, text boxes and tracked changes; headers/footers; comments;
  footnotes/endnotes), plus the package parts (docProps, comment authors,
  ``word/people.xml``) via :mod:`noirdoc.file_analysis.docx_parts`
* **XLSX** – openpyxl cell values plus every part-level surface enumerated by
  :mod:`noirdoc.file_analysis.xlsx_parts` (docProps, comments, headers/footers,
  pivot caches and table captions, chart caches/titles, hyperlinks, filter and
  validation criteria)
* **Plain text** (TXT/CSV/MD/HTML) – simple string replacement

Returns ``None`` for unsupported formats (PDF, PPTX, images) so the
caller can fall through to returning the original bytes.
"""

from __future__ import annotations

import io
import re
from typing import TYPE_CHECKING

import structlog

from noirdoc.mappings.hydration import hydrate_mapper
from noirdoc.reidentification.engine import ReidentificationEngine

if TYPE_CHECKING:
    from noirdoc.pseudonymization.mapper import PseudonymMapper

logger = structlog.get_logger()

_PSEUDO_PATTERN = re.compile(r"<<[A-Z_]+_\d+>>")

_TEXT_MIMES = {
    "text/plain",
    "text/csv",
    "text/markdown",
    "text/html",
    "text/tab-separated-values",
    "application/json",
}

_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def reidentify_file_bytes(
    file_bytes: bytes,
    content_type: str,
    mappings: dict[str, str],
) -> bytes | None:
    """Replace pseudonyms with originals in file bytes.

    Args:
        file_bytes: Raw file content from the provider.
        content_type: MIME type (may include charset parameter).
        mappings: ``{pseudonym: original}`` dict from MappingStore.

    Returns:
        Reidentified bytes, or ``None`` if format is unsupported.
    """
    if not mappings:
        return None

    mapper = hydrate_mapper(mappings)
    engine = ReidentificationEngine()

    # Normalise content_type: strip parameters like "; charset=utf-8"
    mime = content_type.split(";")[0].strip().lower()

    if mime in _TEXT_MIMES:
        return _reidentify_text(file_bytes, engine, mapper)
    if mime == _DOCX_MIME:
        return _reidentify_docx(file_bytes, engine, mapper)
    if mime == _XLSX_MIME:
        return _reidentify_xlsx(file_bytes, engine, mapper)

    return None


def _reidentify_text(
    file_bytes: bytes, engine: ReidentificationEngine, mapper: PseudonymMapper
) -> bytes:
    """Decode → reidentify → encode."""
    text = file_bytes.decode("utf-8", errors="replace")
    reidentified = engine.reidentify(text, mapper)
    return reidentified.encode("utf-8")


def _reidentify_docx(
    file_bytes: bytes, engine: ReidentificationEngine, mapper: PseudonymMapper
) -> bytes | None:
    """Reidentify pseudonyms across every DOCX text surface and package part.

    Both walkers are the ones the redact side writes through, so neither pair
    can drift apart: :mod:`noirdoc.file_analysis.docx_surfaces` for the text
    (body incl. content controls, nested tables, text boxes and tracked
    changes; headers/footers; comments; footnotes/endnotes), and
    :mod:`noirdoc.file_analysis.docx_parts` for the package parts (docProps,
    comment authors/initials, ``word/people.xml``).
    """
    from docx import Document

    from noirdoc.file_analysis.docx_parts import reidentify_document_parts
    from noirdoc.file_analysis.docx_surfaces import rewrite_document_texts

    try:
        doc = Document(io.BytesIO(file_bytes))
    except Exception:
        logger.warning("file_reident.docx_load_failed")
        return None

    def transform(text: str) -> str:
        if not _PSEUDO_PATTERN.search(text):
            return text
        return engine.reidentify(text, mapper)

    changed = rewrite_document_texts(doc, transform, strip_deleted=False)
    if reidentify_document_parts(doc, engine, mapper):
        changed = True

    if not changed:
        return file_bytes

    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _reidentify_xlsx(
    file_bytes: bytes, engine: ReidentificationEngine, mapper: PseudonymMapper
) -> bytes | None:
    """Walk XLSX cells and part-level slots, reidentify string values."""
    from openpyxl import load_workbook

    from noirdoc.file_analysis.xlsx_parts import reidentify_workbook_parts

    try:
        wb = load_workbook(io.BytesIO(file_bytes))
    except Exception:
        logger.warning("file_reident.xlsx_load_failed")
        return None

    changed = False

    for ws in wb.worksheets:  # chartsheets have no cells; their headers are parts
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and _PSEUDO_PATTERN.search(cell.value):
                    new_val = engine.reidentify(cell.value, mapper)
                    if new_val != cell.value:
                        cell.value = new_val
                        changed = True

    # Parts outside the cell grid (docProps, comments, headers/footers, pivot caches) — the
    # same walker the redact side uses, so the two can never drift apart.
    if reidentify_workbook_parts(wb, engine, mapper):
        changed = True

    if not changed:
        wb.close()
        return file_bytes

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
