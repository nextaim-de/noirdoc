"""Smart XLSX pseudonymization using column-type inference.

Three-tier approach for the cell grid:
1. **Header match** – classify columns by header keywords (instant, no NLP)
2. **Sample detection** – run NLP on first N data rows for unclassified columns
3. **Skip** – columns with no PII in header or sample are ignored entirely

Only cells in classified columns are pseudonymized, using direct
``mapper.get_or_create()`` calls instead of full NLP per cell.

Row 1 is the exception. The tiers read it as a header and the rewrite starts at
row 2, so a sheet whose data begins at row 1 — a pasted list, an export with no
header — used to ship its whole first record, silently, while the run still
reported a healthy entity count. Row 1 is now re-examined once the columns are
classified; see :func:`_first_row_candidates` for the rule that decides label
from data without eating real headers.

Everything *outside* the cell grid — ``docProps`` metadata, comments,
headers/footers, pivot caches — is handled by :mod:`noirdoc.file_analysis.xlsx_parts`
after the sheet pass, sharing the same mapper so a name gets one placeholder
wherever it appears.
"""

from __future__ import annotations

import asyncio
import io
from collections.abc import Hashable, Mapping, Sequence
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any, Protocol

import structlog

if TYPE_CHECKING:
    from noirdoc.detection.base import DetectedEntity
    from noirdoc.pseudonymization.mapper import PseudonymMapper

logger = structlog.get_logger()

# Cap on distinct free-text parts (comments, docProps, headers/footers) scanned
# by NER per workbook in detect/block modes. Redact mode ignores the cap: a
# capped redact pass would silently forward unmasked text upstream.
DEFAULT_MAX_FREE_TEXTS = 1000


class XlsxLoadError(RuntimeError):
    """The XLSX package could not be safely loaded for analysis.

    Raised for zip-safety rejections (oversized archives, zip bombs), corrupt
    or truncated packages, and openpyxl parse failures (e.g. pivot caches with
    manually grouped fields). Callers must fail closed: an unloadable workbook
    was never analysed, so its original bytes must not be written out or
    forwarded as if they were redacted.
    """


class DetectorLike(Protocol):
    """Minimal detector interface used here: just async ``detect``.

    Accepts both :class:`~noirdoc.detection.base.BaseDetector` subclasses and
    the structurally-compatible :class:`~noirdoc.detection.ensemble.EnsembleDetector`.
    """

    async def detect(self, text: str, language: str = ...) -> list[DetectedEntity]: ...


# Header keywords → entity type (substring match on normalized header)
_HEADER_ENTITY_MAP: dict[str, str] = {
    # PERSON
    "name": "PERSON",
    "nachname": "PERSON",
    "vorname": "PERSON",
    "firstname": "PERSON",
    "lastname": "PERSON",
    "full name": "PERSON",
    "patient": "PERSON",
    "mitarbeiter": "PERSON",
    "mandant": "PERSON",
    "kunde": "PERSON",
    "klient": "PERSON",
    "bewohner": "PERSON",
    "ansprechpartner": "PERSON",
    "kontaktperson": "PERSON",
    "sachbearbeiter": "PERSON",
    "betreuer": "PERSON",
    "empfänger": "PERSON",
    "absender": "PERSON",
    # EMAIL
    "email": "EMAIL",
    "e-mail": "EMAIL",
    "mail": "EMAIL",
    # PHONE
    "telefon": "PHONE",
    "tel": "PHONE",
    "phone": "PHONE",
    "handy": "PHONE",
    "mobil": "PHONE",
    "fax": "PHONE",
    "rufnummer": "PHONE",
    "durchwahl": "PHONE",
    # LOCATION
    "adresse": "LOCATION",
    "address": "LOCATION",
    "anschrift": "LOCATION",
    "wohnort": "LOCATION",
    "straße": "LOCATION",
    "strasse": "LOCATION",
    "ort": "LOCATION",
    "stadt": "LOCATION",
    "plz": "LOCATION",
    "postleitzahl": "LOCATION",
    # DATE
    "geburtsdatum": "DATE",
    "geburtstag": "DATE",
    "datum": "DATE",
    "date": "DATE",
    "birthday": "DATE",
    "geb": "DATE",
    # IBAN
    "iban": "IBAN",
    "kontonummer": "IBAN",
    "bankverbindung": "IBAN",
    # SVNR
    "sozialversicherungsnummer": "SVNR",
    "svnr": "SVNR",
    "sv-nummer": "SVNR",
    "rentenversicherungsnummer": "SVNR",
    "versicherungsnummer": "SVNR",
    # STEUER_ID
    "steuer-id": "STEUER_ID",
    "steuerid": "STEUER_ID",
    "steueridentifikationsnummer": "STEUER_ID",
    "steuernummer": "STEUER_ID",
    "identifikationsnummer": "STEUER_ID",
    "tin": "STEUER_ID",
    "idnr": "STEUER_ID",
    # ORGANIZATION
    "firma": "ORGANIZATION",
    "unternehmen": "ORGANIZATION",
    "company": "ORGANIZATION",
    "arbeitgeber": "ORGANIZATION",
    "auftraggeber": "ORGANIZATION",
    "kanzlei": "ORGANIZATION",
}


def infer_entity_type(header_value: object) -> str | None:
    """Match a header cell value against the keyword map (substring match)."""
    if not isinstance(header_value, str) or not header_value.strip():
        return None
    header_lower = header_value.strip().lower()
    for keyword, entity_type in _HEADER_ENTITY_MAP.items():
        if keyword in header_lower:
            return entity_type
    return None


async def detect_each(
    texts: Sequence[str],
    detector: DetectorLike,
    language: str,
    *,
    concurrency: int = 8,
) -> list[list[DetectedEntity]]:
    """Run the detector over *texts* concurrently, results in the same order."""
    if not texts:
        return []

    sem = asyncio.Semaphore(concurrency)

    async def _detect(text: str) -> list[DetectedEntity]:
        async with sem:
            return await detector.detect(text, language)

    return await asyncio.gather(*[_detect(text) for text in texts])


async def classify_by_sample[K: Hashable](
    samples: Sequence[tuple[K, str]],
    detector: DetectorLike,
    language: str,
    *,
    concurrency: int = 8,
) -> dict[K, str]:
    """Tier 2: classify keys (columns, pivot fields, …) from sampled cell texts.

    Samples are detected concurrently; for each key the first sample (in the
    given order) that yields any entity decides, using the highest-scoring
    entity's type. Keys without a hit are absent from the result.
    """
    if not samples:
        return {}

    results = await detect_each(
        [text for _, text in samples], detector, language, concurrency=concurrency
    )

    classified: dict[K, str] = {}
    for (key, _), entities in zip(samples, results, strict=True):
        if entities and key not in classified:
            best = max(entities, key=lambda e: e.score)
            classified[key] = best.entity_type
    return classified


async def _detect_first_row(
    candidates: Sequence[Any],
    detector: DetectorLike,
    language: str,
) -> dict[int, list[DetectedEntity]]:
    """Detect over the row-1 cells that might be data rather than a label.

    Each distinct string is detected once — a row 1 repeated across columns
    ("n/a", a date) costs one call, not one per column.
    """
    if not candidates:
        return {}
    order: dict[str, int] = {}
    for cell in candidates:
        order.setdefault(cell.value, len(order))
    results = await detect_each(list(order), detector, language)
    return {cell.column: results[order[cell.value]] for cell in candidates}


def _first_row_candidates(
    header_row: Sequence[Any],
    keyword_cols: set[int],
    col_types: Mapping[int, str | None],
    *,
    single_row: bool,
) -> list[Any]:
    """Row-1 cells that may hold data instead of a column label.

    Tier 3 starts at row 2, so nothing else in the sheet pass ever looks at row
    1: a list pasted without a header, or an export that starts at row 1, used
    to ship its whole first record. Two things keep this from eating real
    headers, which would be its own kind of damage:

    * A cell whose own text matched the keyword map is a label by definition
      ("Name", "E-Mail") — never a candidate.
    * Otherwise the column has to have been classified from rows 2 onwards, so
      row 1 is only rewritten where the data below it says that column holds
      PII. A header the keyword map missed ("Betrag") sits over numbers, gets
      no classification, and is left alone even if NER would flag the word.

    A single-row sheet has no rows below to compare against, and no reason to
    hold a header with nothing under it, so every cell is a candidate. Those
    sheets used to be skipped outright.
    """
    return [
        cell
        for cell in header_row
        if cell.column not in keyword_cols
        and isinstance(cell.value, str)
        and cell.value.strip()
        and (single_row or col_types.get(cell.column) not in (None, "skip"))
    ]


def _pseudonymize_first_row(
    header_row: Sequence[Any],
    row_one_entities: Mapping[int, list[DetectedEntity]],
    mapper: PseudonymMapper,
    result: XlsxResult,
    *,
    apply: bool,
) -> None:
    """Replace the detected spans in row 1, span-wise like free text.

    Span-wise rather than whole-cell (what tier 3 does for a classified column)
    because row 1 is not known to be a bare value: "Kontakt Anna Mueller" keeps
    its prefix. The mapper is shared, so a cell that *is* just the name gets the
    same placeholder the same name gets in row 5.
    """
    from noirdoc.pseudonymization.engine import PseudonymizationEngine

    engine = PseudonymizationEngine()
    for cell in header_row:
        entities = row_one_entities.get(cell.column)
        if not entities or not isinstance(cell.value, str):
            continue
        if apply:
            cell.value = engine.pseudonymize(cell.value, entities, mapper)
        for entity in entities:
            result.entity_count += 1
            result.entity_types[entity.entity_type] = (
                result.entity_types.get(entity.entity_type, 0) + 1
            )
        best = max(entities, key=lambda e: e.score)
        # Keyed by position, never by the cell's text — that text is the PII.
        result.column_classifications[f"row1!col{cell.column}"] = f"{best.entity_type} (row 1 data)"


@dataclass
class XlsxResult:
    """Result of smart XLSX pseudonymization.

    ``entity_count`` / ``entity_types`` cover cells *and* part-level hits
    (metadata, comments, headers/footers, pivot caches). They also include PII
    in parts the writer drops rather than maps (``app.xml`` Manager/Company,
    ``xl/persons``) — those are reported as ``"… (dropped)"`` in
    ``column_classifications`` and are not reversible.
    """

    new_bytes: bytes | None = None
    entity_count: int = 0
    entity_types: dict[str, int] = field(default_factory=dict)
    column_classifications: dict[str, str] = field(default_factory=dict)
    # Distinct free-text parts (comments, docProps, headers/footers) NOT run
    # through the detector — cap or block-mode early exit, detect/block only.
    free_texts_skipped: int = 0

    def merge_parts(self, entity_types: dict[str, int], classifications: dict[str, str]) -> None:
        for entity_type, count in entity_types.items():
            self.entity_types[entity_type] = self.entity_types.get(entity_type, 0) + count
            self.entity_count += count
        self.column_classifications.update(classifications)


async def pseudonymize_xlsx_smart(
    data: bytes,
    detector: DetectorLike,
    mapper: PseudonymMapper,
    language: str = "de",
    sample_rows: int = 5,
    pseudonymize: bool = True,
    max_free_texts: int | None = DEFAULT_MAX_FREE_TEXTS,
    stop_on_first_hit: bool = False,
) -> XlsxResult:
    """Analyse and optionally pseudonymize an XLSX file using column-type inference.

    When *pseudonymize* is ``True``, cells in classified columns are replaced
    via ``mapper.get_or_create()`` and the modified workbook is returned.
    When ``False``, cells are only counted (for detect-only / block modes).

    *max_free_texts* (``None`` = no cap) and *stop_on_first_hit*
    (block mode: any hit settles the decision, so the free-text NER pass stops
    early) bound the part-level scan. Both are ignored when *pseudonymize* is
    ``True`` — redact mode must scan everything it rewrites. Skipped texts are
    reported in ``XlsxResult.free_texts_skipped``, never dropped silently.

    Raises :class:`XlsxLoadError` when the package cannot be loaded (zip-safety
    rejection, corrupt archive, openpyxl parse failure). Returning an empty
    result here would let callers treat the original workbook as redacted.
    """
    from openpyxl import load_workbook

    from noirdoc.file_analysis.extractors._zip_safety import check_ooxml_zip_safe

    result = XlsxResult()

    try:
        check_ooxml_zip_safe(data, label="xlsx")
        wb = load_workbook(io.BytesIO(data))
    except Exception as exc:
        logger.warning("xlsx_inference.load_failed", error=str(exc))
        raise XlsxLoadError(f"cannot load xlsx workbook: {exc}") from exc

    # Lazy import: xlsx_parts imports infer_entity_type / classify_by_sample from here.
    from noirdoc.file_analysis.xlsx_parts import (
        clear_phantom_creator,
        count_unsupported_part_pii,
        pseudonymize_workbook_parts,
    )

    stop_on_first_hit = stop_on_first_hit and not pseudonymize

    clear_phantom_creator(wb, data)
    # lower-cased header -> entity type, so pivot fields built from a classified column
    # are treated the same way as the column itself.
    known_fields: dict[str, str] = {}

    # Chartsheets have no cell grid; their headers/footers are handled with the parts.
    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 1:
            continue

        # --- Tier 1: classify columns from header row ---
        col_types: dict[int, str | None] = {}
        header_row = list(next(ws.iter_rows(min_row=1, max_row=1)))
        keyword_cols: set[int] = set()
        for cell in header_row:
            etype = infer_entity_type(cell.value)
            col_types[cell.column] = etype
            if etype:
                keyword_cols.add(cell.column)
                label = cell.value if isinstance(cell.value, str) else f"col{cell.column}"
                result.column_classifications[label] = f"{etype} (header)"

        # --- Tier 2: sample first N data rows for unclassified columns ---
        unknown_cols = {col for col, t in col_types.items() if t is None}
        sampled: dict[int, str] = {}
        if unknown_cols:
            sample_cells: list[tuple[int, str]] = []
            for row in ws.iter_rows(min_row=2, max_row=min(1 + sample_rows, ws.max_row)):
                for cell in row:
                    if (
                        cell.column in unknown_cols
                        and isinstance(cell.value, str)
                        and cell.value.strip()
                    ):
                        sample_cells.append((cell.column, cell.value))

            sampled = await classify_by_sample(sample_cells, detector, language)
            for col_idx, sampled_type in sampled.items():
                col_types[col_idx] = sampled_type

            # Mark remaining unknowns as skip
            for col in unknown_cols:
                if col_types[col] is None:
                    col_types[col] = "skip"

        # --- Row 1: a header label, or the first record? ---
        row_one_entities = await _detect_first_row(
            _first_row_candidates(header_row, keyword_cols, col_types, single_row=ws.max_row == 1),
            detector,
            language,
        )
        data_cols = {col for col, entities in row_one_entities.items() if entities}

        # Labels for the sampled columns are assigned here rather than above,
        # because a row-1 cell that turned out to be data must not become a
        # dict key: column_classifications is logged.
        for col_idx, sampled_type in sampled.items():
            hcell = next((c for c in header_row if c.column == col_idx), None)
            label = (
                hcell.value
                if hcell and isinstance(hcell.value, str) and col_idx not in data_cols
                else f"col{col_idx}"
            )
            result.column_classifications[label] = f"{sampled_type} (sampled)"

        for cell in header_row:
            col_type = col_types.get(cell.column)
            if (
                isinstance(cell.value, str)
                and col_type
                and col_type != "skip"
                and cell.column not in data_cols  # a value, not a field name
            ):
                known_fields.setdefault(cell.value.strip().lower(), col_type)

        # Must run after known_fields, which reads the original row-1 strings.
        _pseudonymize_first_row(header_row, row_one_entities, mapper, result, apply=pseudonymize)

        # --- Tier 3: process all data rows for classified columns ---
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                entity_type = col_types.get(cell.column)
                if not entity_type or entity_type == "skip":
                    continue
                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue
                if pseudonymize:
                    cell.value = mapper.get_or_create(cell.value, entity_type)
                result.entity_count += 1
                result.entity_types[entity_type] = result.entity_types.get(entity_type, 0) + 1

    # --- Parts outside the cell grid: docProps, comments, headers/footers, pivot caches ---
    # Block mode: a cell-grid hit already settles the decision, so skip the
    # free-text NER pass entirely (reported as skipped, not silently dropped).
    parts_free_limit = 0 if stop_on_first_hit and result.entity_count > 0 else max_free_texts
    parts = await pseudonymize_workbook_parts(
        wb,
        detector,
        mapper,
        language,
        apply=pseudonymize,
        sample_size=sample_rows,
        known_fields=known_fields,
        max_free_texts=parts_free_limit,
        stop_on_first_hit=stop_on_first_hit,
    )
    result.merge_parts(parts.entity_types, parts.classifications)
    result.free_texts_skipped = parts.free_texts_skipped
    dropped = count_unsupported_part_pii(data)
    result.merge_parts(dropped.entity_types, dropped.classifications)

    if pseudonymize and result.entity_count > 0:
        buf = io.BytesIO()
        wb.save(buf)
        result.new_bytes = buf.getvalue()

    wb.close()

    logger.info(
        "xlsx_inference.completed",
        entity_count=result.entity_count,
        entity_types=result.entity_types,
        columns=result.column_classifications,
        free_texts_skipped=result.free_texts_skipped,
    )

    return result
