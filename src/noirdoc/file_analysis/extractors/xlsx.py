"""XLSX text extraction using openpyxl.

This helper flattens an entire workbook into a single string and is intended
for the "ship XLSX as text to a non-Excel-aware LLM" path
(``file_analysis.pipeline.convert_unsupported_files``). It is **not** suitable
for redaction — the joined text destroys cell context. Redaction must go
through :func:`noirdoc.file_analysis.xlsx_inference.pseudonymize_xlsx_smart`,
which preserves columns and writes per-cell pseudonyms.
"""

from __future__ import annotations

import io

from noirdoc.file_analysis.extractors._zip_safety import check_ooxml_zip_safe


def extract_xlsx(data: bytes) -> str:
    """Extract cell values from all sheets of an XLSX byte-string."""
    from openpyxl import load_workbook

    check_ooxml_zip_safe(data, label="xlsx")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(v) for v in row if v is not None)
                if row_text.strip():
                    lines.append(row_text)
    finally:
        wb.close()
    return "\n".join(lines)
